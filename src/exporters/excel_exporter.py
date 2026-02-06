"""
Excel exporter for liquidation documents.

Creates well-formatted Excel workbooks with multiple sheets for different data sections.
"""
import pandas as pd
from pathlib import Path
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.models.liquidation import LiquidationDocument


class ExcelExporter:
    """Export liquidation documents to Excel with formatting."""

    def __init__(self, document: LiquidationDocument):
        self.document = document

    def export(self, output_path: str):
        """
        Export document to Excel file.

        Args:
            output_path: Path where Excel file will be saved
        """
        output_path = Path(output_path)

        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write different sections to different sheets
            self._write_document_info(writer)
            self._write_tribute_records(writer)
            self._write_exercise_summaries(writer)

        # Apply formatting
        self._apply_formatting(output_path)

    def _write_document_info(self, writer):
        """Write document header information."""
        info_data = {
            'Campo': [
                'Ejercicio',
                'Entidad',
                'Total Registros',
                'Total C_CARGO',
                'Total C_DATAS',
                'Total C_VOLUNTARIA',
                'Total C_EJECUTIVA',
                'Total CC_PENDIENTE',
                'Total C_TOTAL'
            ],
            'Valor': [
                self.document.ejercicio,
                self.document.entidad,
                self.document.total_records,
                float(self.document.total_c_cargo),
                float(self.document.total_c_datas),
                float(self.document.total_c_voluntaria),
                float(self.document.total_c_ejecutiva),
                float(self.document.total_cc_pendiente),
                float(self.document.total_c_total)
            ]
        }

        df = pd.DataFrame(info_data)
        df.to_excel(writer, sheet_name='Información', index=False)

    def _write_tribute_records(self, writer):
        """Write tribute records table."""
        records_data = []

        for record in self.document.tribute_records:
            records_data.append({
                'C_EJERCICIO': record.ejercicio,
                'C_CONCEPTO': record.concepto,
                'CLAVE_C': record.clave_contabilidad,
                'CLAVE_R': record.clave_recaudacion,
                'C_CARGO': float(record.c_cargo),
                'C_DATAS': float(record.c_datas),
                'C_VOLUNTARIA': float(record.c_voluntaria),
                'C_EJECUTIVA': float(record.c_ejecutiva),
                'CC_PENDIENTE': float(record.cc_pendiente),
                'C_TOTAL': float(record.c_total)
            })

        df = pd.DataFrame(records_data)
        df.to_excel(writer, sheet_name='Registros', index=False)

    def _write_exercise_summaries(self, writer):
        """Write exercise summaries."""
        summaries_data = []

        for summary in self.document.exercise_summaries:
            summaries_data.append({
                'Ejercicio': summary.ejercicio,
                'C_CARGO': float(summary.c_cargo),
                'C_DATAS': float(summary.c_datas),
                'C_VOLUNTARIA': float(summary.c_voluntaria),
                'C_EJECUTIVA': float(summary.c_ejecutiva),
                'CC_PENDIENTE': float(summary.cc_pendiente),
                'C_TOTAL': float(summary.c_total),
                'Número de Registros': len(summary.records)
            })

        # Add overall totals
        summaries_data.append({
            'Ejercicio': 'TOTAL',
            'C_CARGO': float(self.document.total_c_cargo),
            'C_DATAS': float(self.document.total_c_datas),
            'C_VOLUNTARIA': float(self.document.total_c_voluntaria),
            'C_EJECUTIVA': float(self.document.total_c_ejecutiva),
            'CC_PENDIENTE': float(self.document.total_cc_pendiente),
            'C_TOTAL': float(self.document.total_c_total),
            'Número de Registros': self.document.total_records
        })

        df = pd.DataFrame(summaries_data)
        df.to_excel(writer, sheet_name='Resumen por Ejercicio', index=False)


    def _apply_formatting(self, file_path: Path):
        """Apply formatting to the Excel workbook."""
        wb = load_workbook(file_path)

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_font = Font(bold=True, size=11)

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Format header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Format number columns
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    # Apply border to all cells
                    cell.border = border

                    # Format numbers
                    if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

            # Highlight total rows in "Resumen por Ejercicio"
            if sheet_name == 'Resumen por Ejercicio':
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == 'TOTAL':
                        for cell in row:
                            cell.fill = total_fill
                            cell.font = total_font

            # Freeze first row
            ws.freeze_panes = ws['A2']

        wb.save(file_path)


def export_to_excel(document: LiquidationDocument, output_path: str):
    """
    Convenience function to export a liquidation document to Excel.

    Args:
        document: The LiquidationDocument to export
        output_path: Path where Excel file will be saved
    """
    exporter = ExcelExporter(document)
    exporter.export(output_path)
